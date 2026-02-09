"""
generar_f29.py v5 — Genera Excel F29 con estructura completa del SII.

Replica exactamente la estructura, colores y fórmulas del formulario oficial F29
(194 filas, 150 líneas, todas las secciones).

Uso:
    from scripts.generar_f29 import generar_f29_excel
    generar_f29_excel(datos, output_path)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ============================================================
# Estilos — Colores del F29 oficial
# ============================================================
FG = PatternFill(start_color="73B464", end_color="73B464", fill_type="solid")  # Green
FB = PatternFill(start_color="D9EDF7", end_color="D9EDF7", fill_type="solid")  # Blue
FE = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")  # Gray
FL = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")  # Light gray
FW = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

F14B = Font(name="Arial", size=14, bold=True)
F11BW = Font(name="Arial", size=11, bold=True, color="FFFFFF")
F11 = Font(name="Arial", size=11)
F10 = Font(name="Arial", size=10)
F9B = Font(name="Arial", size=9, bold=True)
F8B = Font(name="Arial", size=8, bold=True)
F8 = Font(name="Arial", size=8)
F7C = Font(name="Arial", size=7, color="666666")

AL = Alignment(horizontal="left", vertical="center", wrap_text=True)
AC = Alignment(horizontal="center", vertical="center", wrap_text=True)
AR = Alignment(horizontal="right", vertical="center", wrap_text=True)

NF = '#,##0'
PF = '0.0#'
NCOLS = 7

MESES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
}


def formato_peso(v):
    """Formatea un valor como peso chileno ($1.234.567)."""
    if v is None or v == 0:
        return "$0"
    if isinstance(v, float):
        v = int(v)
    if v < 0:
        return f"-${abs(v):,}".replace(",", ".")
    return f"${v:,}".replace(",", ".")

# ============================================================
# Datos de líneas del F29
# Formato: (num, descripción, cod_cant, cod_monto, operador)
# ============================================================

L_DEB_INFO = [
    (1, 'Exportaciones', 585, 20, None),
    (2, 'Ventas y/o Servicios prestados Exentos, o No Gravados del giro', 586, 142, None),
    (3, 'Ventas con retención sobre el margen de comercialización (contribuyentes retenidos)', 731, 732, None),
    (4, 'Ventas y/o Servicios prestados exentos o No Gravados que no son del giro', 714, 715, None),
    (5, 'Facturas de Compra recibidas con retención total (contribuyentes retenidos) y Factura de Inicio emitida', 515, 587, None),
    (6, 'Facturas de compras recibidas con retención parcial (Total neto)', None, 720, None),
]

L_DEB_GENERA = [
    (7, 'Facturas emitidas por ventas y servicios del giro, o por cuenta de terceros', 503, 502, '+'),
    (8, 'Facturas emitidas por la venta de bienes inmuebles afectas a IVA', 763, 764, '+'),
    (9, 'Facturas y Notas de Débitos por ventas y servicios que no son del giro (activo fijo y otros)', 716, 717, '+'),
    (10, 'Boletas', 110, 111, '+'),
    (11, 'Comprobantes o Recibos de Pago generados en transacciones pagadas a través de medios electrónicos', 758, 759, '+'),
    (12, 'Notas de Débito emitidas asociadas al giro y ND recibidas de terceros por retención parcial de cambio de sujeto', 512, 513, '+'),
    (13, 'Notas de Crédito emitidas por Facturas asociadas al giro y NC recibidas de terceros por retención parcial de cambio de sujeto', 509, 510, '-'),
    (14, 'Notas de Crédito emitidas por Vales de máquinas autorizadas por el Servicio', 708, 709, '-'),
    (15, 'Notas de Crédito emitidas por ventas y servicios que no son del giro (activo fijo y otros)', 733, 734, '-'),
    (16, 'Facturas de Compra recibidas con retención parcial (contribuyentes retenidos)', 516, 517, '+'),
    (17, 'Liquidación y Liquidación Factura', 500, 501, '+'),
    (18, 'Adiciones al Débito Fiscal del mes por Art.27 bis', None, 154, '+'),
    (19, 'Restitución Adicional Art.27 bis, inc.2º (Ley 19.738)', None, 518, '+'),
    (20, 'Reintegro Impuesto de Timbres y Estampillas, Art. 3º Ley Nº 20.259 e IVA Arrendamiento esporádico BBII Amoblados', None, 713, '+'),
    (21, 'Adiciones al Débito por IEPD Ley 20.765', 738, 741, '+'),
    (22, 'Restitución Adicional Reembolso Remanente CF IVA (Ley 21.256)', None, 791, '+'),
]

L_CRED_SIN = [
    (25, 'Internas afectas', 564, 521, None),
    (26, 'Importaciones', 566, 560, None),
    (27, 'Internas exentas, o no gravadas', 584, 562, None),
]

L_CRED_INT = [
    (28, 'Facturas recibidas del giro y Facturas de compra emitidas', 519, 520, '+'),
    (29, 'Facturas recibidas de Proveedores: Supermercados y Comercios similares, Art.23 Nº4 D.L.825 (Ley Nº20.780)', 761, 762, '+'),
    (30, 'Facturas recibidas por Adquisición o Construcción de Bienes Inmuebles, Art.8º transitorio (Ley Nº20.780)', 765, 766, '+'),
    (31, 'Facturas activo fijo', 524, 525, '+'),
    (32, 'Notas de Crédito recibidas y NC emitidas por retención de cambio de sujeto', 527, 528, '-'),
    (33, 'Notas de Débito recibidas y ND emitidas por retención de cambio de sujeto', 531, 532, '+'),
]

L_CRED_IMP = [
    (34, 'Declaraciones de Ingreso (DIN) importaciones del giro', 534, 535, '+'),
    (35, 'Declaraciones de Ingreso (DIN) importaciones de activo fijo', 536, 553, '+'),
]

L_CRED_REM = [
    (36, 'Remanente Crédito Fiscal mes anterior', None, 504, '+'),
    (37, 'Devolución Solicitud Art. 36 (Exportadores)', None, 593, '-'),
    (38, 'Devolución Solicitud Art. 27 bis (Activo fijo)', None, 594, '-'),
    (39, 'Certificado Imputación Art. 27 bis (Activo fijo)', None, 592, '-'),
    (40, 'Devolución Solicitud Art. 3º (Cambio de Sujeto)', None, 539, '-'),
    (41, 'Devolución Solicitud Ley Nº 20.258 (Generadoras Eléctricas)', None, 718, '-'),
    (42, 'Devolución Solicitud Reembolso Remanente de Crédito Fiscal IVA', None, 790, '-'),
    (43, 'Monto Reintegrado por Devolución Indebida de Crédito Fiscal D.S. 348 (Exportadores)', None, 164, '+'),
]

L_CRED_IEPD = [
    (44, 'Recuperación de Impuesto Específico al Petróleo Diesel (Art. 7º Ley 18.502, Arts.1º y 3º D.S. Nº 311/86)', 730, 127, '+'),
    (45, 'Recuperación Imp. Específico Petróleo Diesel Transportistas de Carga (Art. 2º Ley Nº19.764)', 729, 544, '+'),
]

L_CRED_OTROS = [
    (46, 'Crédito del Art.11º Ley 18.211 (Zona Franca de Extensión)', None, 523, '+'),
    (47, 'Crédito por Impuesto de Timbres y Estampillas, Art. 3º Ley Nº 20.259', None, 712, '+'),
    (48, 'Crédito por IVA restituido a aportantes sin domicilio ni residencia en Chile (Art. 83, Ley 20.712)', None, 757, '+'),
]

L_POST_51 = [(51, 'Saldo de IVA postergado en 12 cuotas', 772, 775, '+')]

L_POST_CUOTAS = [
    (52, 'Monto Total de IVA postergado', 777, 780, '+'),
    (53, 'Monto total IVA postergado (Ley 20.780)', 782, 783, '+'),
    (54, 'Monto total IVA postergado (Ley 21.207)', 784, 785, '+'),
    (55, 'Monto Total IVA postergado (DIN)', 786, 787, '+'),
    (56, 'Monto Total IVA postergado (Tributación Simplificada)', 788, 789, '+'),
    (57, 'Restitución de devolución Art. 27 ter D.L. 825, inc. 2º (Ley Nº 20.720)', None, 760, '+'),
    (58, 'Certificado Imputación Art. 27 ter D.L. 825, inc. 1º (Ley Nº 20.720)', None, 767, '-'),
]

L_RET = [
    (59, 'Retención Imp. Primera Categoría por rentas de capitales mobiliarios del Art.20 Nº2, según Art.73 LIR', None, 50, '+'),
    (60, 'Retención Impuesto Único a los Trabajadores, según Art. 74 Nº 1 LIR', 751, 48, '+'),
    (61, 'Retención de Impuesto con tasa del 15.25% sobre las rentas del Art. 42 Nº2, según Art. 74 Nº2 LIR', None, 151, '+'),
    (62, 'Retención de Impuesto con tasa del 10% sobre las rentas del Art. 48, según Art. 74 N° 3 LIR', None, 153, '+'),
    (63, 'Retención sobre rentas del Art. 42 N°1 LIR con tasa del 3%', None, 49, '+'),
    (64, 'Retención sobre rentas del Art. 42 N°2 LIR con tasa del 3%', None, 155, '+'),
    (65, 'Retención a Suplementeros, según Art. 74 N°5 (tasa 0,5%) LIR', None, 54, '+'),
    (66, 'Retención por compra de productos mineros, según Art. 74 N° 6 LIR', None, 56, '+'),
    (67, 'Retención sobre rescates y seguros dotales del N° 3 del Art.17 LIR (tasa 15%)', None, 588, '+'),
    (68, 'Retención sobre retiros de Ahorro Previsional Voluntario del Art. 42 bis LIR (tasa 15%)', None, 589, '+'),
]

L_PPM = [
    (69, '1ra Categoría Art. 84 a) y 14 D N° 3 letra (k) y 8 letra (a) numeral (viii)', 750, 62, '+'),
    (70, '1ra Cat. Art. 84 a) tasa 3%, reintegro préstamo tasa 0%', None, 156, '+'),
    (71, 'Mineros, Art.84 a)', 565, 123, '+'),
    (72, 'Explotador Minero Art. 84 h)', 700, 703, '+'),
    (73, 'Explotador Minero Royalty Ley 21.591', 806, 810, '+'),
    (74, 'Transportistas acogidos a Renta Presunta, Art 84, e) y f) (tasa de 0,3%)', None, 66, '+'),
    (75, 'Crédito Capacitación, Ley 19.518/97', 721, 723, '-'),
    (76, '2da. Categoría Art. 84, b) (tasa 15.25%)', None, 152, '+'),
    (77, '2da. Cat. Art. 84 b) LIR tasa 3%', None, 157, '+'),
    (78, 'Taller artesanal Art.84, c) (tasa de 1,5% o 3%)', None, 70, '+'),
    (79, 'Renta Líquida Provisional inciso final de la letra a) del art 84 de la LIR, Ley N° 21.210', None, 776, '+'),
]

L_TRIB_SIMP = [
    (81, 'Ventas del período', None, 529, None),
    (82, 'Crédito del período', None, 530, None),
    (83, 'IVA determinado por concepto de Tributación Simplificada', None, 409, '+'),
]

L_ART37 = [
    (84, 'Letras e), h), i), l) (tasa 15%)', None, 522, '+'),
    (85, 'Letra j) (tasa 50%)', None, 526, '+'),
    (86, 'Débito de Impuesto Adicional Ventas Art. 37 letras a), b) y c) y Art. 40 D.L.825 (tasa 15%)', None, 113, '+'),
    (87, 'Crédito de Impuesto Adicional Art.37 letras a), b) y c) D.L. 825', None, 28, '-'),
    (88, 'Monto reintegrado por devolución indebida de crédito por exportadores D.L. 825', None, 548, '-'),
    (89, 'Remanente crédito Art. 37 mes anterior D.L.825', None, 540, '-'),
    (90, 'Devolución Solicitud Art.36 relativa al Imp. Adicional Art.37 letras a), b) y c) D.L. 825', None, 541, '+'),
]

L_ART42_DEB = [
    (92, 'Pisco, Licores, Whisky y Aguardiente (tasa 31,5%)', None, 577, '+'),
    (93, 'Vinos, Champaña, Chichas (tasa 20,5%)', None, 32, '+'),
    (94, 'Cervezas (tasa 20,5%)', None, 150, '+'),
    (95, 'Bebidas analcohólicas (tasa 10%)', None, 146, '+'),
    (96, 'Bebidas analcohólicas elevado contenido azúcares (tasa 18%)', None, 752, '+'),
    (97, 'Notas de Débito emitidas', None, 545, '+'),
    (98, 'Notas de Crédito emitidas por Facturas', None, 546, '-'),
    (99, 'Notas de Crédito emitidas por Vales de máquinas autorizadas por el Servicio', None, 710, '-'),
]

L_ART42_CRED = [
    (101, 'Pisco, Licores, Whisky y Aguardiente (tasa 31,5%)', 575, 576, '+'),
    (102, 'Vinos, Champaña, Chichas (tasa 20,5%)', 574, 33, '+'),
    (103, 'Cervezas (tasa 20,5%)', 580, 149, '+'),
    (104, 'Bebidas analcohólicas (tasa 10%)', 582, 85, '+'),
    (105, 'Bebidas analcohólicas elevado contenido azúcares (tasa 18%)', 753, 754, '+'),
    (106, 'Notas de Débito recibidas', None, 551, '+'),
    (107, 'Notas de Crédito recibidas', None, 559, '-'),
    (108, 'Remanente crédito Art.42 mes anterior', None, 508, '+'),
    (109, 'Devolución Art. 36 D.L.825 relativas impuesto Art.42', None, 533, '-'),
    (110, 'Monto reintegrado devoluciones indebidas de crédito por exportaciones', None, 552, '+'),
]

L_ANTICIPO_CS = [
    (113, 'IVA anticipado del período', None, 556, '+'),
    (114, 'Remanente del mes anterior', None, 557, '+'),
    (115, 'Devolución del mes anterior', None, 558, '-'),
]

L_CS_AGENTE = [
    (118, 'IVA total retenido a terceros (tasa Art. 14 DL 825)', None, 39, '+'),
    (119, 'IVA parcial retenido a terceros (según tasa)', None, 554, '+'),
    (120, 'IVA Retenido por notas de crédito emitidas', None, 736, '-'),
    (121, 'Retención de margen de comercialización', None, 597, '+'),
    (122, 'Retención Anticipo de Cambio de Sujeto', 555, 596, '+'),
]

L_CS_ESPECIAL = [
    (123, 'IVA retenido a terceros con retención total en el período', None, 100, '+'),
    (124, 'Ajustes por concepto de IVA asociado a reversiones y contracargos (disputas) solucionadas en el período', None, 101, '-'),
    (125, 'Valor nominal del remanente de ajuste (código 104 del período anterior)', None, 102, '-'),
]

L_VENTA_REMOTA = [
    (128, 'IVA total del periodo por la venta remota de bienes corporales muebles', None, 811, '+'),
    (129, 'Ajustes por concepto de IVA asociado a reversiones y contracargos solucionados en el período', None, 812, '-'),
    (130, 'Valor nominal del remanente de ajuste (código [815] del período anterior)', None, 813, '-'),
]

L_CRED_ESP = [
    (134, 'Crédito por Sistemas Solares Térmicos, Ley 20.365', 725, 727, '-'),
    (135, 'Imputación del Pago Patente Aguas Ley 20.017', 704, 706, '-'),
    (136, 'Cotización Adicional Ley 18.566', 160, 570, '-'),
    (137, 'Crédito Especial Empresas Constructoras', 126, 571, '-'),
    (138, 'Recup. Peajes Transportistas Pasajeros, Ley 19.764', 572, 590, '-'),
    (139, 'Crédito por desembolsos directos trazabilidad', 768, 770, '-'),
]

L_REM_CRED_ESP = [
    (141, 'Remanente Crédito por Sistemas Solares Térmicos, Ley 20.365', None, 728, None),
    (142, 'Remanente periodo siguiente Patente Aguas, Ley 20.017', None, 707, None),
    (143, 'Remanente de Cotización Adicional Ley 18.566', None, 73, None),
    (144, 'Remanente Crédito Especial Empresas Constructoras', None, 130, None),
    (145, 'Remanente Recup. de Peajes Trans. Pasajeros Ley 19.764', None, 591, None),
    (146, 'Remanente Crédito por desembolsos directos trazabilidad', None, 771, None),
]

# Todas las líneas que generan crédito (para fórmula línea 49)
ALL_CRED_LINES = L_CRED_INT + L_CRED_IMP + L_CRED_REM + L_CRED_IEPD + L_CRED_OTROS

# Todas las líneas 50-79 para fórmula línea 80 (con sus operadores)
# Se construye dinámicamente en _write_f29 porque incluye línea 50 especial


# ============================================================
# Helpers de escritura
# ============================================================

def _g(ws, r, text):
    """Green section header (merged)."""
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS)
    ws.cell(row=r, column=1, value=text).font = F11BW
    ws.cell(row=r, column=1).alignment = AC
    for c in range(1, NCOLS + 1):
        ws.cell(row=r, column=c).fill = FG
    return r + 1


def _b(ws, r, text):
    """Blue sub-section header (merged, left-aligned)."""
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS)
    ws.cell(row=r, column=1, value=text).font = F9B
    ws.cell(row=r, column=1).alignment = AL
    for c in range(1, NCOLS + 1):
        ws.cell(row=r, column=c).fill = FB
    return r + 1


def _h(ws, r, d_label='Cantidad de Documentos', f_label='Monto Neto',
       g_label='+/-', a_label='Línea', b_label='Descripción'):
    """Column headers row (blue fill)."""
    for c in range(1, NCOLS + 1):
        ws.cell(row=r, column=c).fill = FB
        ws.cell(row=r, column=c).font = F8B
        ws.cell(row=r, column=c).alignment = AC
    if a_label:
        ws.cell(row=r, column=1, value=a_label)
    if b_label:
        ws.cell(row=r, column=2, value=b_label)
    ws.cell(row=r, column=3, value='Cód')
    if d_label:
        ws.cell(row=r, column=4, value=d_label)
    ws.cell(row=r, column=5, value='Cód')
    if f_label:
        ws.cell(row=r, column=6, value=f_label)
    if g_label:
        ws.cell(row=r, column=7, value=g_label)
    return r + 1


def _h2(ws, r, b_label, d_label, f_label, g_label='+/-'):
    """Header row with custom B column label and no A/C labels."""
    for c in range(1, NCOLS + 1):
        ws.cell(row=r, column=c).fill = FB
        ws.cell(row=r, column=c).font = F8B
        ws.cell(row=r, column=c).alignment = AC
    if b_label:
        ws.cell(row=r, column=2, value=b_label)
    if d_label:
        ws.cell(row=r, column=4, value=d_label)
    ws.cell(row=r, column=5, value='Cód')
    if f_label:
        ws.cell(row=r, column=6, value=f_label)
    if g_label:
        ws.cell(row=r, column=7, value=g_label)
    return r + 1


def _ln(ws, r, num, desc, cq, ca, op, cod, cc, bold=False):
    """Standard data line."""
    # Col A: line number
    a = ws.cell(row=r, column=1, value=num)
    a.font = F8; a.fill = FE; a.alignment = AC
    # Col B: description
    ws.cell(row=r, column=2, value=desc).font = F8B if bold else F8
    ws.cell(row=r, column=2).alignment = AL
    # Col C + D: qty code + value
    if cq is not None:
        ws.cell(row=r, column=3, value=cq).font = F7C
        ws.cell(row=r, column=3).fill = FE
        ws.cell(row=r, column=3).alignment = AC
        cc[cq] = f"D{r}"
        v = cod.get(cq)
        if v:
            d = ws.cell(row=r, column=4, value=v)
            d.font = F11; d.alignment = AR; d.number_format = NF
    # Col E + F: amt code + value
    if ca is not None:
        ws.cell(row=r, column=5, value=ca).font = F7C
        ws.cell(row=r, column=5).fill = FE
        ws.cell(row=r, column=5).alignment = AC
        cc[ca] = f"F{r}"
        v = cod.get(ca)
        if v:
            f = ws.cell(row=r, column=6, value=v)
            f.font = F11; f.alignment = AR; f.number_format = NF
    # Col G: operator
    if op:
        ws.cell(row=r, column=7, value=op).font = F9B
        ws.cell(row=r, column=7).alignment = AC
    return r + 1


def _wl(ws, r, lines, cod, cc):
    """Write a list of data lines."""
    for num, desc, cq, ca, op in lines:
        r = _ln(ws, r, num, desc, cq, ca, op, cod, cc)
    return r


def _fv(ws, r, col, formula):
    """Write a formula value cell (LGRAY fill, 11pt, right, #,##0)."""
    c = ws.cell(row=r, column=col, value=formula)
    c.font = F11; c.fill = FL; c.alignment = AR; c.number_format = NF
    return c


def _fl(ws, r, num, desc, ca, formula, op, cc, bold=True):
    """Write a formula total line (no qty code)."""
    ws.cell(row=r, column=1, value=num).font = F8
    ws.cell(row=r, column=1).fill = FE
    ws.cell(row=r, column=1).alignment = AC
    ws.cell(row=r, column=2, value=desc).font = F8B if bold else F8
    ws.cell(row=r, column=2).alignment = AL
    if ca is not None:
        ws.cell(row=r, column=5, value=ca).font = F7C
        ws.cell(row=r, column=5).fill = FE
        ws.cell(row=r, column=5).alignment = AC
        cc[ca] = f"F{r}"
    _fv(ws, r, 6, formula)
    if op:
        ws.cell(row=r, column=7, value=op).font = F9B
        ws.cell(row=r, column=7).alignment = AC
    return r + 1


def _bf(lines, cc):
    """Build formula string from lines based on their +/- operators."""
    parts = []
    for num, desc, cq, ca, op in lines:
        ref = cc.get(ca)
        if ref and op in ('+', '-'):
            parts.append(f"{op}{ref}")
    if not parts:
        return "=0"
    s = "=" + "".join(parts)
    if s.startswith("=+"):
        s = "=" + s[2:]
    return s


# ============================================================
# Cálculo del F29
# ============================================================

def calcular_f29(datos):
    """Calcula códigos del F29 desde datos."""
    if "codigos" in datos:
        codigos = dict(datos["codigos"])
        for k in [538, 537, 89, 77, 595, 547, 91, 62]:
            codigos.setdefault(k, 0)
        return codigos

    v = datos.get("ventas", {})
    c = datos.get("compras", {})
    ret = datos.get("retenciones", {})
    ppm = datos.get("ppm", {})
    dev = datos.get("devoluciones", {})
    docs = datos.get("documentos", {})
    cs = datos.get("cambio_sujeto", {})
    IVA_TASA = 0.19
    codigos = {}

    def has_docs(lk): return lk in docs and len(docs[lk]) > 0
    def count_docs(lk): return len(docs.get(lk, []))
    def sum_field(lk, campo): return sum(d.get(campo, 0) for d in docs.get(lk, []))

    # DÉBITOS
    for cq, ca, lk, fc, fn, is_iva in [
        (585, 20, "linea_1", "facturas_exportacion_cant", "facturas_exportacion_neto", False),
        (586, 142, "linea_2", "facturas_exentas_giro_cant", "facturas_exentas_giro_neto", False),
        (515, 587, "linea_5", "facturas_compra_digital_cant", "facturas_compra_digital_neto", False),
        (503, 502, "linea_7", "facturas_afectas_cant", "facturas_afectas_neto", True),
        (716, 717, "linea_9", "ventas_activo_fijo_cant", "ventas_activo_fijo_neto", True),
        (110, 111, "linea_10", "boletas_cant", "boletas_neto", True),
        (512, 513, "linea_12", "notas_debito_cant", "notas_debito_neto", True),
        (509, 510, "linea_13", "notas_credito_cant", "notas_credito_neto", True),
    ]:
        if has_docs(lk):
            codigos[cq] = count_docs(lk)
            codigos[ca] = sum_field(lk, "iva") if is_iva else sum_field(lk, "neto")
            if is_iva and not codigos[ca]:
                codigos[ca] = int(sum_field(lk, "neto") * IVA_TASA)
        else:
            codigos[cq] = v.get(fc, c.get(fc, 0))
            codigos[ca] = int(v.get(fn, 0) * IVA_TASA) if is_iva else v.get(fn, c.get(fn, 0))

    for k in [731, 732, 714, 715, 720, 763, 764, 758, 759, 708, 709, 733, 734,
              516, 517, 500, 501, 154, 518, 713, 738, 741, 791]:
        codigos.setdefault(k, 0)

    codigos[538] = sum(codigos.get(ca, 0) * (1 if op == '+' else -1)
                       for _, _, _, ca, op in L_DEB_GENERA if op in ('+', '-'))

    # CRÉDITOS
    for cq, ca, lk, fc, fi in [
        (519, 520, "linea_28", "facturas_giro_cant", "facturas_giro_iva"),
        (524, 525, "linea_31", "facturas_activo_fijo_cant", "facturas_activo_fijo_iva"),
        (527, 528, "linea_32", "notas_credito_recibidas_cant", "notas_credito_recibidas_iva"),
        (531, 532, "linea_33", "notas_debito_recibidas_cant", "notas_debito_recibidas_iva"),
        (534, 535, "linea_34", "din_giro_cant", "din_giro_iva"),
        (536, 553, "linea_35", "din_activo_fijo_cant", "din_activo_fijo_iva"),
    ]:
        if has_docs(lk):
            codigos[cq] = count_docs(lk)
            codigos[ca] = sum_field(lk, "iva")
        else:
            codigos[cq] = c.get(fc, 0)
            codigos[ca] = c.get(fi, 0)

    for k in [761, 762, 765, 766, 564, 521, 566, 560, 730, 127, 729, 544]:
        codigos.setdefault(k, 0)
    codigos.setdefault(584, c.get("exentas_sin_derecho_cant", 0))
    codigos.setdefault(562, c.get("exentas_sin_derecho_neto", 0))
    codigos[511] = codigos.get(519, 0) + codigos.get(524, 0)
    codigos[514] = codigos.get(520, 0) + codigos.get(525, 0)
    codigos[504] = datos.get("remanente_anterior", 0)
    codigos[593] = dev.get("art_36_exportador", 0)
    codigos[594] = dev.get("art_27_bis", 0)
    codigos[592] = dev.get("certificado_27_bis", 0)
    codigos[539] = dev.get("cambio_sujeto", 0)
    for k in [718, 790, 164, 523, 712, 757]:
        codigos.setdefault(k, 0)

    codigos[537] = sum(codigos.get(ca, 0) * (1 if op == '+' else -1)
                       for _, _, _, ca, op in ALL_CRED_LINES if op in ('+', '-'))

    # IVA DETERMINADO
    td, tc = codigos[538], codigos[537]
    codigos[89] = max(td - tc, 0)
    codigos[77] = max(tc - td, 0)

    # RETENCIONES
    codigos.setdefault(50, 0)
    if has_docs("linea_60"):
        codigos[48] = sum(d.get("iusc", d.get("iva", 0)) for d in docs["linea_60"])
    else:
        codigos[48] = ret.get("iusc_impuesto", 0)
    if has_docs("linea_61"):
        codigos[151] = sum(d.get("retencion", d.get("iva", 0)) for d in docs["linea_61"])
    else:
        codigos[151] = ret.get("honorarios_retencion", 0)
    codigos[153] = ret.get("directores_retencion", 0)
    for k in [49, 155, 54, 56, 588, 589, 751]:
        codigos.setdefault(k, 0)

    # PPM
    tasa_ppm = ppm.get("tasa", 0.25)
    codigos[115] = tasa_ppm
    codigos.setdefault(750, 0)
    codigos.setdefault(30, 0)
    if ppm.get("base_imponible") is not None:
        base_ppm = ppm["base_imponible"]
    else:
        def neto_linea(lk, fk):
            return sum_field(lk, "neto") if has_docs(lk) else v.get(fk, 0)
        base_ppm = (neto_linea("linea_7", "facturas_afectas_neto")
                    + neto_linea("linea_2", "facturas_exentas_giro_neto")
                    + neto_linea("linea_1", "facturas_exportacion_neto")
                    + neto_linea("linea_10", "boletas_neto")
                    + neto_linea("linea_12", "notas_debito_neto")
                    - neto_linea("linea_13", "notas_credito_neto"))
    codigos[563] = base_ppm
    codigos[68] = 0
    codigos[62] = 0 if ppm.get("suspension") else int(base_ppm * tasa_ppm / 100)
    codigos[722] = ppm.get("remanente_sence_anterior", 0)
    codigos[721] = ppm.get("credito_sence", 0)
    sence = codigos[721] + codigos[722]
    codigos[723] = min(sence, codigos[62])
    codigos[724] = sence - codigos[723]

    # SUB TOTAL (línea 80)
    total_ret = sum(codigos.get(ca, 0) for _, _, _, ca, _ in L_RET)
    ppm_neto = codigos[62] - codigos[723]
    codigos[595] = codigos[89] + total_ret + ppm_neto

    # CAMBIO DE SUJETO
    codigos[39] = cs.get("iva_retenido_total", cs.get(39, 0))
    codigos[554] = cs.get("iva_parcial_retenido", cs.get(554, 0))
    codigos[736] = cs.get("iva_retenido_nc", cs.get(736, 0))
    codigos[597] = cs.get("retencion_margen", cs.get(597, 0))
    codigos[596] = cs.get("retencion_neta", cs.get(596, 0))
    if codigos[596] == 0 and codigos[39] > 0:
        codigos[596] = codigos[39] + codigos.get(554, 0) - codigos[736] + codigos.get(597, 0)

    # TOTAL
    codigos[547] = codigos[595] + codigos.get(596, 0)
    codigos[91] = codigos[547]
    codigos[92] = 0; codigos[93] = 0
    codigos[94] = codigos[91]
    return codigos


# ============================================================
# Hoja 1: Formulario F29
# ============================================================

def _write_f29(wb, codigos, enc):
    """Hoja 1: F29 completo (194 filas) con fórmulas."""
    mes = enc.get("periodo_mes", 1)
    anio = enc.get("periodo_anio", 2026)
    ws = wb.active
    ws.title = f"F29 — {MESES.get(mes)} {anio}"

    for letter, width in [('A', 6), ('B', 55), ('C', 7), ('D', 18), ('E', 7), ('F', 18), ('G', 4)]:
        ws.column_dimensions[letter].width = width

    cod = codigos
    cc = {}  # code -> cell ref
    r = 1

    # --- TÍTULO Y ENCABEZADO ---
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS)
    ws.cell(row=r, column=1, value='DECLARACIÓN MENSUAL Y PAGO SIMULTÁNEO DE IMPUESTOS - FORMULARIO 29').font = F14B
    ws.cell(row=r, column=1).alignment = AC
    r += 1

    rut = enc.get("rut", "____________")
    folio = enc.get("folio", "____________")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS)
    ws.cell(row=r, column=1, value=f'Período Tributario: {mes:02d}/{anio}    RUT: {rut}    Folio: {folio}').font = F10
    ws.cell(row=r, column=1).alignment = AC
    for c in range(1, NCOLS + 1):
        ws.cell(row=r, column=c).fill = FB
    r += 1

    # --- DÉBITOS Y VENTAS ---
    r = _g(ws, r, 'DÉBITOS Y VENTAS')
    r = _b(ws, r, 'VENTAS Y/O SERVICIOS PRESTADOS')
    r = _b(ws, r, 'INFORMACIÓN DE INGRESOS')
    r = _h(ws, r, 'Cantidad de Documentos', 'Monto Neto', g_label=None)
    r = _wl(ws, r, L_DEB_INFO, cod, cc)

    r = _h(ws, r, 'Cantidad de Documentos', 'Débito', b_label='GENERA DÉBITO')
    r = _wl(ws, r, L_DEB_GENERA, cod, cc)

    # Línea 23: TOTAL DÉBITOS
    r = _fl(ws, r, 23, 'TOTAL DÉBITOS', 538, _bf(L_DEB_GENERA, cc), '=', cc, bold=True)

    # --- CRÉDITOS Y COMPRAS ---
    r = _g(ws, r, 'CRÉDITOS Y COMPRAS')
    r = _b(ws, r, 'COMPRAS Y/O SERVICIOS UTILIZADOS')
    r = _h(ws, r, 'Con Derecho a Crédito', 'Sin Derecho a Crédito', g_label=None)
    r = _ln(ws, r, 24, 'IVA por documentos electrónicos recibidos', 511, 514, None, cod, cc)

    r = _b(ws, r, 'SIN DERECHO A CRÉDITO FISCAL')
    r = _h(ws, r, 'Cantidad de Documentos', 'Monto Neto', g_label=None)
    r = _wl(ws, r, L_CRED_SIN, cod, cc)

    r = _b(ws, r, 'CON DERECHO A CRÉDITO FISCAL')
    r = _b(ws, r, 'INTERNAS')
    r = _h(ws, r, 'Cantidad de Documentos', 'Crédito, Recuperación y Reintegro')
    r = _wl(ws, r, L_CRED_INT, cod, cc)

    r = _b(ws, r, 'IMPORTACIONES')
    r = _wl(ws, r, L_CRED_IMP, cod, cc)
    r = _wl(ws, r, L_CRED_REM, cod, cc)

    r = _b(ws, r, 'LEY 20.765')
    r = _h(ws, r, 'M3 Comprados con Derecho a Crédito', 'Componentes del Impuesto')
    r = _wl(ws, r, L_CRED_IEPD, cod, cc)
    r = _wl(ws, r, L_CRED_OTROS, cod, cc)

    # Línea 49: TOTAL CRÉDITOS
    r = _fl(ws, r, 49, 'TOTAL CRÉDITOS', 537, _bf(ALL_CRED_LINES, cc), '=', cc, bold=True)

    # --- POSTERGACIÓN DE IVA ---
    r = _h2(ws, r, 'POSTERGACIÓN DE IVA', 'Remanente CF', 'Impuesto Determinado')

    # Línea 50: ESPECIAL — dual codes (77 en C/D, 89 en E/F)
    ref_538 = cc.get(538, '0')
    ref_537 = cc.get(537, '0')
    ws.cell(row=r, column=1, value=50).font = F8
    ws.cell(row=r, column=1).fill = FE; ws.cell(row=r, column=1).alignment = AC
    ws.cell(row=r, column=2, value='Remanente de crédito fiscal para el período siguiente').font = F8
    ws.cell(row=r, column=2).alignment = AL
    ws.cell(row=r, column=3, value=77).font = F7C
    ws.cell(row=r, column=3).fill = FE; ws.cell(row=r, column=3).alignment = AC
    cc[77] = f"D{r}"
    _fv(ws, r, 4, f'=IF({ref_537}>{ref_538},{ref_537}-{ref_538},0)')
    ws.cell(row=r, column=5, value=89).font = F7C
    ws.cell(row=r, column=5).fill = FE; ws.cell(row=r, column=5).alignment = AC
    cc[89] = f"F{r}"
    _fv(ws, r, 6, f'=IF({ref_538}>{ref_537},{ref_538}-{ref_537},0)')
    ws.cell(row=r, column=7, value='+').font = F9B
    ws.cell(row=r, column=7).alignment = AC
    r += 1

    # Línea 51
    r = _wl(ws, r, L_POST_51, cod, cc)

    # Postergación IVA en cuotas
    # Header especial con C:D merged
    for c in range(1, NCOLS + 1):
        ws.cell(row=r, column=c).fill = FB
        ws.cell(row=r, column=c).font = F8B
        ws.cell(row=r, column=c).alignment = AC
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    ws.cell(row=r, column=2, value='POSTERGACIÓN IVA EN CUOTAS (D.S. 420/997 MH)')
    ws.cell(row=r, column=5, value='Cód')
    ws.cell(row=r, column=6, value='Imp. Determinado')
    ws.cell(row=r, column=7, value='+/-')
    r += 1
    r = _wl(ws, r, L_POST_CUOTAS, cod, cc)

    # --- IMPUESTO A LA RENTA ---
    r = _g(ws, r, 'IMPUESTO A LA RENTA D.L. 824/74')
    r = _b(ws, r, 'RETENCIONES')
    r = _h(ws, r, d_label=None, f_label='Impuesto Determinado', a_label='Línea', b_label='Descripción')
    r = _wl(ws, r, L_RET, cod, cc)

    # PPM
    r = _b(ws, r, 'PPM')
    r = _h(ws, r, 'Base Imponible / Tasa', 'PPM Determinado')
    # Para línea 69: escribir base imponible en D (code 563 interno)
    for num, desc, cq, ca, op in L_PPM:
        if num == 69:
            # Línea especial: D = base imponible, F = PPM
            ws.cell(row=r, column=1, value=69).font = F8
            ws.cell(row=r, column=1).fill = FE; ws.cell(row=r, column=1).alignment = AC
            ws.cell(row=r, column=2, value=desc).font = F8
            ws.cell(row=r, column=2).alignment = AL
            ws.cell(row=r, column=3, value=750).font = F7C
            ws.cell(row=r, column=3).fill = FE; ws.cell(row=r, column=3).alignment = AC
            cc[750] = f"D{r}"
            # Escribir base imponible en D
            base = cod.get(563, 0)
            if base:
                d = ws.cell(row=r, column=4, value=base)
                d.font = F11; d.alignment = AR; d.number_format = NF
            ws.cell(row=r, column=5, value=62).font = F7C
            ws.cell(row=r, column=5).fill = FE; ws.cell(row=r, column=5).alignment = AC
            cc[62] = f"F{r}"
            v = cod.get(62, 0)
            if v:
                f = ws.cell(row=r, column=6, value=v)
                f.font = F11; f.alignment = AR; f.number_format = NF
            ws.cell(row=r, column=7, value='+').font = F9B
            ws.cell(row=r, column=7).alignment = AC
            r += 1
        else:
            r = _ln(ws, r, num, desc, cq, ca, op, cod, cc)

    # Línea 80: SUB TOTAL
    # Fórmula dinámica: IVA det + postergación + retenciones + PPM (con signos)
    line80_parts = []
    # IVA determinado (línea 50)
    ref_89 = cc.get(89)
    if ref_89:
        line80_parts.append(ref_89)
    # Línea 51
    for _, _, _, ca, op in L_POST_51:
        ref = cc.get(ca)
        if ref and op in ('+', '-'):
            line80_parts.append(f"{op}{ref}")
    # Líneas 52-58
    for _, _, _, ca, op in L_POST_CUOTAS:
        ref = cc.get(ca)
        if ref and op in ('+', '-'):
            line80_parts.append(f"{op}{ref}")
    # Retenciones 59-68
    for _, _, _, ca, op in L_RET:
        ref = cc.get(ca)
        if ref and op in ('+', '-'):
            line80_parts.append(f"{op}{ref}")
    # PPM 69-79
    for _, _, _, ca, op in L_PPM:
        ref = cc.get(ca)
        if ref and op in ('+', '-'):
            line80_parts.append(f"{op}{ref}")
    formula_80 = "=" + "".join(line80_parts)
    if formula_80.startswith("=+"):
        formula_80 = "=" + formula_80[2:]
    r = _fl(ws, r, 80, 'Sub total impuesto determinado anverso (Suma de las líneas 50 a 78)',
            595, formula_80, '=', cc, bold=True)

    # --- TRIBUTACIÓN SIMPLIFICADA ---
    r = _g(ws, r, 'SISTEMA DE TRIBUTACIÓN SIMPLIFICADA DEL IVA, ART. 29 D.L. 825')
    r = _h(ws, r, d_label=None, f_label='Impuesto Determinado')
    r = _wl(ws, r, L_TRIB_SIMP, cod, cc)

    # --- IMPUESTO ADICIONAL ART. 37 ---
    r = _g(ws, r, 'IMPUESTO ADICIONAL ART. 37 D.L. 825')
    r = _h(ws, r, d_label=None, f_label='Impuesto Determinado')
    r = _wl(ws, r, L_ART37, cod, cc)

    # Línea 91: dual codes (549/D=remanente, 550/F=impuesto)
    art37_formula = _bf(L_ART37, cc)
    ws.cell(row=r, column=1, value=91).font = F8
    ws.cell(row=r, column=1).fill = FE; ws.cell(row=r, column=1).alignment = AC
    ws.cell(row=r, column=2, value='Remanente crédito impuesto Art.37 para período siguiente').font = F8
    ws.cell(row=r, column=2).alignment = AL
    ws.cell(row=r, column=3, value=549).font = F7C
    ws.cell(row=r, column=3).fill = FE; ws.cell(row=r, column=3).alignment = AC
    cc[549] = f"D{r}"
    # D = IF(sum>=0, sum, 0) — impuesto
    inner = art37_formula[1:]  # remove leading =
    _fv(ws, r, 4, f'=IF(({inner})>=0,{inner},0)')
    ws.cell(row=r, column=5, value=550).font = F7C
    ws.cell(row=r, column=5).fill = FE; ws.cell(row=r, column=5).alignment = AC
    cc[550] = f"F{r}"
    # F = IF(sum<0, ABS(sum), 0) — remanente
    _fv(ws, r, 6, f'=IF(({inner})<0,ABS({inner}),0)')
    ws.cell(row=r, column=7, value='+').font = F9B
    ws.cell(row=r, column=7).alignment = AC
    r += 1

    # --- IMPUESTO ADICIONAL ART. 42 ---
    r = _g(ws, r, 'IMPUESTO ADICIONAL ART. 42 D.L. 825')
    r = _h2(ws, r, 'DÉBITOS', None, 'Débito')
    r = _wl(ws, r, L_ART42_DEB, cod, cc)

    # Línea 100: Total Débitos Art 42
    r = _fl(ws, r, 100, 'Total Débitos Art. 42 DL 825', 602, _bf(L_ART42_DEB, cc), '=', cc, bold=True)

    r = _h2(ws, r, 'CRÉDITOS', 'Total Crédito Recargado Facturas Recibidas', 'Crédito Imputable del Periodo')
    r = _wl(ws, r, L_ART42_CRED, cod, cc)

    # Línea 111: Total Créditos Art 42
    r = _fl(ws, r, 111, 'Total créditos Art.42 DL 825', 603, _bf(L_ART42_CRED, cc), '=', cc, bold=True)

    # Línea 112: dual codes (507/D=remanente, 506/F=impuesto)
    ref_602 = cc.get(602, '0')
    ref_603 = cc.get(603, '0')
    ws.cell(row=r, column=1, value=112).font = F8
    ws.cell(row=r, column=1).fill = FE; ws.cell(row=r, column=1).alignment = AC
    ws.cell(row=r, column=2, value='Remanente crédito Imp. Adic. Art.42 para período siguiente').font = F8
    ws.cell(row=r, column=2).alignment = AL
    ws.cell(row=r, column=3, value=507).font = F7C
    ws.cell(row=r, column=3).fill = FE; ws.cell(row=r, column=3).alignment = AC
    cc[507] = f"D{r}"
    _fv(ws, r, 4, f'=IF({ref_602}>{ref_603},{ref_602}-{ref_603},0)')
    ws.cell(row=r, column=5, value=506).font = F7C
    ws.cell(row=r, column=5).fill = FE; ws.cell(row=r, column=5).alignment = AC
    cc[506] = f"F{r}"
    _fv(ws, r, 6, f'=IF({ref_603}>{ref_602},{ref_603}-{ref_602},0)')
    ws.cell(row=r, column=7, value='+').font = F9B
    ws.cell(row=r, column=7).alignment = AC
    r += 1

    # --- CAMBIO DE SUJETO ---
    r = _g(ws, r, 'CAMBIO DE SUJETO D.L. 825')
    r = _b(ws, r, 'ANTICIPO CAMBIO DE SUJETO (CONTRIBUYENTES RETENIDOS)')
    r = _h(ws, r, d_label=None, f_label='Monto')
    r = _wl(ws, r, L_ANTICIPO_CS, cod, cc)

    # Línea 116: Total Anticipo
    r = _fl(ws, r, 116, 'Total de Anticipo', 543, _bf(L_ANTICIPO_CS, cc), '=', cc, bold=True)

    # Línea 117: Remanente Anticipos
    ws.cell(row=r, column=1, value=117).font = F8
    ws.cell(row=r, column=1).fill = FE; ws.cell(row=r, column=1).alignment = AC
    ws.cell(row=r, column=2, value='Remanente Anticipos Cambio Sujeto para período siguiente').font = F8
    ws.cell(row=r, column=2).alignment = AL
    ws.cell(row=r, column=3, value=573).font = F7C
    ws.cell(row=r, column=3).fill = FE; ws.cell(row=r, column=3).alignment = AC
    cc[573] = f"D{r}"
    ws.cell(row=r, column=5, value=598).font = F7C
    ws.cell(row=r, column=5).fill = FE; ws.cell(row=r, column=5).alignment = AC
    cc[598] = f"F{r}"
    ref_543 = cc.get(543, '0')
    _fv(ws, r, 6, f'=IF({ref_543}<0,ABS({ref_543}),0)')
    ws.cell(row=r, column=7, value='-').font = F9B
    ws.cell(row=r, column=7).alignment = AC
    r += 1

    # CAMBIO DE SUJETO (AGENTE RETENEDOR)
    r = _b(ws, r, 'CAMBIO DE SUJETO (AGENTE RETENEDOR)')
    r = _wl(ws, r, L_CS_AGENTE, cod, cc)

    # CAMBIO ESPECIAL DE SUJETO
    r = _b(ws, r, 'CAMBIO ESPECIAL DE SUJETO (Inciso 7º, Art. 3º D.L. 825)')
    r = _wl(ws, r, L_CS_ESPECIAL, cod, cc)

    # Línea 126: Monto neto IVA retenido
    r = _fl(ws, r, 126, 'Monto neto de IVA retenido en el período', 103, _bf(L_CS_ESPECIAL, cc), '=', cc, bold=True)

    # Línea 127: Remanente ajuste
    ws.cell(row=r, column=1, value=127).font = F8
    ws.cell(row=r, column=1).fill = FE; ws.cell(row=r, column=1).alignment = AC
    ws.cell(row=r, column=2, value='Remanente de ajuste para el próximo período').font = F8
    ws.cell(row=r, column=2).alignment = AL
    ws.cell(row=r, column=5, value=104).font = F7C
    ws.cell(row=r, column=5).fill = FE; ws.cell(row=r, column=5).alignment = AC
    cc[104] = f"F{r}"
    ref_103 = cc.get(103, '0')
    _fv(ws, r, 6, f'=IF({ref_103}<0,ABS({ref_103}),0)')
    ws.cell(row=r, column=7, value='=').font = F9B
    ws.cell(row=r, column=7).alignment = AC
    r += 1

    # IVA POR VENTA REMOTA
    r = _b(ws, r, 'IVA POR LA VENTA REMOTA DE BIENES CORPORALES MUEBLES (Art. 3° bis e inciso final del art. 4°, D.L. 825)')
    r = _wl(ws, r, L_VENTA_REMOTA, cod, cc)

    # Línea 131: Monto neto venta remota
    r = _fl(ws, r, 131, 'Monto neto de IVA del período', 814, _bf(L_VENTA_REMOTA, cc), '=', cc, bold=True)

    # Línea 132: Remanente venta remota
    ws.cell(row=r, column=1, value=132).font = F8
    ws.cell(row=r, column=1).fill = FE; ws.cell(row=r, column=1).alignment = AC
    ws.cell(row=r, column=2, value='Remanente de ajuste para el próximo período').font = F8
    ws.cell(row=r, column=2).alignment = AL
    ws.cell(row=r, column=5, value=815).font = F7C
    ws.cell(row=r, column=5).fill = FE; ws.cell(row=r, column=5).alignment = AC
    cc[815] = f"F{r}"
    ref_814 = cc.get(814, '0')
    _fv(ws, r, 6, f'=IF({ref_814}<0,ABS({ref_814}),0)')
    ws.cell(row=r, column=7, value='=').font = F9B
    ws.cell(row=r, column=7).alignment = AC
    r += 1

    # IMPUESTO SUSTITUTIVO
    r = _b(ws, r, 'IMPUESTO SUSTITUTIVO RETENIDO POR RÉGIMEN TRIBUTARIO ESPECIAL A COMERCIANTES DE FERIAS LIBRES')
    r = _ln(ws, r, 133, 'Impuesto sustitutivo retenido por régimen tributario especial a comerciantes de ferias libres',
            None, 816, '=', cod, cc)

    # --- CRÉDITOS ESPECIALES ---
    r = _g(ws, r, 'CRÉDITOS ESPECIALES')
    r = _h(ws, r, 'Base / Remanente Anterior', 'Crédito del Período')
    r = _wl(ws, r, L_CRED_ESP, cod, cc)

    # --- LÍNEA 140: TOTAL DETERMINADO ---
    # Fórmula compleja: F105 + F110 + D120 + D144 + IF(F151>0,F151,0) + ...
    ref_595 = cc.get(595, '0')
    ref_409 = cc.get(409, '0')
    ref_549 = cc.get(549, '0')
    ref_507 = cc.get(507, '0')
    ref_543_2 = cc.get(543, '0')
    parts_140 = [ref_595, f"+{ref_409}", f"+{ref_549}", f"+{ref_507}"]
    parts_140.append(f"+IF({ref_543_2}>0,{ref_543_2},0)")
    for _, _, _, ca, op in L_CS_AGENTE:
        ref = cc.get(ca)
        if ref and op in ('+', '-'):
            parts_140.append(f"{op}{ref}")
    ref_103_2 = cc.get(103, '0')
    parts_140.append(f"+IF({ref_103_2}>0,{ref_103_2},0)")
    ref_814_2 = cc.get(814, '0')
    parts_140.append(f"+IF({ref_814_2}>0,{ref_814_2},0)")
    ref_816 = cc.get(816, '0')
    parts_140.append(f"+{ref_816}")
    for _, _, _, ca, op in L_CRED_ESP:
        ref = cc.get(ca)
        if ref and op in ('+', '-'):
            parts_140.append(f"{op}{ref}")
    formula_140 = "=" + "".join(parts_140)
    if formula_140.startswith("=+"):
        formula_140 = "=" + formula_140[2:]
    r = _fl(ws, r, 140, 'TOTAL DETERMINADO', 547, formula_140, '=', cc, bold=True)

    # --- REMANENTE CRÉDITOS ESPECIALES ---
    r = _g(ws, r, 'REMANENTE CRÉDITOS ESPECIALES')
    r = _h(ws, r, d_label=None, f_label='Remanente', g_label=None)
    r = _wl(ws, r, L_REM_CRED_ESP, cod, cc)

    # --- TOTAL A PAGAR ---
    ref_547 = cc.get(547, '0')
    r = _fl(ws, r, 147, 'TOTAL A PAGAR EN PLAZO LEGAL', 91, f'={ref_547}', '=', cc, bold=True)
    r = _ln(ws, r, 148, 'Más IPC', None, 92, '+', cod, cc)
    r = _ln(ws, r, 149, 'Más Intereses y multas', None, 93, '+', cod, cc)

    # Condonación
    ws.cell(row=r, column=2, value='Condonación').font = F8
    ws.cell(row=r, column=2).alignment = AL
    ws.cell(row=r, column=5, value=60).font = F7C
    ws.cell(row=r, column=5).fill = FE; ws.cell(row=r, column=5).alignment = AC
    cc[60] = f"F{r}"
    r += 1

    # Línea 150: TOTAL CON RECARGO
    ref_91 = cc.get(91, '0')
    ref_92 = cc.get(92, '0')
    ref_93 = cc.get(93, '0')
    r = _fl(ws, r, 150, 'TOTAL A PAGAR CON RECARGO', 94, f'={ref_91}+{ref_92}+{ref_93}', '=', cc, bold=True)


# ============================================================
# Hoja 2: Detalle de documentos
# ============================================================

LINEAS_INFO = {
    "linea_1":  {"linea": "1",  "cod_cant": 585, "cod_monto": 20,  "nombre": "Facturas de Exportación", "seccion": "debito"},
    "linea_2":  {"linea": "2",  "cod_cant": 586, "cod_monto": 142, "nombre": "Ventas/Servicios Exentos del Giro", "seccion": "debito"},
    "linea_5":  {"linea": "5",  "cod_cant": 515, "cod_monto": 587, "nombre": "Facturas de Compra (Serv. Digitales Extranjeros)", "seccion": "debito"},
    "linea_7":  {"linea": "7",  "cod_cant": 503, "cod_monto": 502, "nombre": "Facturas Afectas del Giro", "seccion": "debito"},
    "linea_9":  {"linea": "9",  "cod_cant": 716, "cod_monto": 717, "nombre": "Ventas Activo Fijo (No del Giro)", "seccion": "debito"},
    "linea_10": {"linea": "10", "cod_cant": 110, "cod_monto": 111, "nombre": "Boletas", "seccion": "debito"},
    "linea_11": {"linea": "11", "cod_cant": 758, "cod_monto": 759, "nombre": "Boletas Electrónicas / POS", "seccion": "debito"},
    "linea_12": {"linea": "12", "cod_cant": 512, "cod_monto": 513, "nombre": "Notas de Débito Emitidas", "seccion": "debito"},
    "linea_13": {"linea": "13", "cod_cant": 509, "cod_monto": 510, "nombre": "Notas de Crédito Emitidas", "seccion": "debito"},
    "linea_27": {"linea": "27", "cod_cant": 584, "cod_monto": 562, "nombre": "Internas Exentas sin Derecho a CF", "seccion": "credito"},
    "linea_28": {"linea": "28", "cod_cant": 519, "cod_monto": 520, "nombre": "Facturas Recibidas del Giro + FC Emitidas", "seccion": "credito"},
    "linea_29": {"linea": "29", "cod_cant": 761, "cod_monto": 762, "nombre": "Facturas Supermercados/Comercios", "seccion": "credito"},
    "linea_31": {"linea": "31", "cod_cant": 524, "cod_monto": 525, "nombre": "Facturas Activo Fijo", "seccion": "credito"},
    "linea_32": {"linea": "32", "cod_cant": 527, "cod_monto": 528, "nombre": "Notas de Crédito Recibidas / NC por Cambio de Sujeto", "seccion": "credito"},
    "linea_33": {"linea": "33", "cod_cant": 531, "cod_monto": 532, "nombre": "Notas de Débito Recibidas", "seccion": "credito"},
    "linea_60": {"linea": "60", "cod_cant": None, "cod_monto": 48,  "nombre": "Impuesto Único 2da Categoría (Sueldos)", "seccion": "retencion"},
    "linea_61": {"linea": "61", "cod_cant": None, "cod_monto": 151, "nombre": "Retención Honorarios Art. 42 N°2", "seccion": "retencion"},
}

COLUMNAS_VENTAS = ["N° Doc", "Fecha", "RUT", "Razón Social", "Descripción", "Neto", "IVA", "Exento", "Total"]
COLUMNAS_COMPRAS = ["N° Doc", "Fecha", "RUT", "Razón Social", "Descripción", "Neto", "IVA", "Total"]
COLUMNAS_HONORARIOS = ["N° Boleta", "Fecha", "RUT", "Razón Social", "Descripción", "Bruto", "Retención", "Líquido"]
COLUMNAS_SUELDOS = ["N° Liquidación", "Fecha", "RUT", "Nombre", "Cargo", "Sueldo Bruto", "IUSC", "Líquido"]
N_TEXT_COLS = 5


def _get_columnas(seccion, lk):
    if lk == "linea_61": return COLUMNAS_HONORARIOS
    if lk == "linea_60": return COLUMNAS_SUELDOS
    if seccion == "debito": return COLUMNAS_VENTAS
    return COLUMNAS_COMPRAS


def _get_doc_values(doc, seccion, lk):
    if lk == "linea_61":
        return [doc.get("numero", ""), doc.get("fecha", ""), doc.get("rut", ""),
                doc.get("razon_social", ""), doc.get("descripcion", ""),
                doc.get("bruto", doc.get("neto", 0)), doc.get("retencion", doc.get("iva", 0)),
                doc.get("liquido", doc.get("total", 0))]
    if lk == "linea_60":
        return [doc.get("numero", ""), doc.get("fecha", ""), doc.get("rut", ""),
                doc.get("razon_social", doc.get("nombre", "")),
                doc.get("cargo", doc.get("descripcion", "")),
                doc.get("bruto", doc.get("neto", 0)), doc.get("iusc", doc.get("iva", 0)),
                doc.get("liquido", doc.get("total", 0))]
    if seccion == "debito":
        return [doc.get("numero", ""), doc.get("fecha", ""), doc.get("rut", ""),
                doc.get("razon_social", ""), doc.get("descripcion", ""),
                doc.get("neto", 0), doc.get("iva", 0), doc.get("exento", 0), doc.get("total", 0)]
    return [doc.get("numero", ""), doc.get("fecha", ""), doc.get("rut", ""),
            doc.get("razon_social", ""), doc.get("descripcion", ""),
            doc.get("neto", 0), doc.get("iva", 0), doc.get("total", 0)]


def _write_detalle(wb, datos, codigos=None):
    """Hoja 2: Detalle de documentos por línea con fórmulas en totales."""
    docs = datos.get("documentos", {})
    if not docs:
        ws = wb.create_sheet(title="Detalle Documentos")
        ws.cell(row=1, column=1, value="Sin documentos individuales.").font = F8B
        return

    ws = wb.create_sheet(title="Detalle Documentos")
    for i, w in enumerate([4, 14, 12, 20, 30, 30, 14, 14, 14, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    c = ws.cell(row=r, column=1, value="DETALLE DE DOCUMENTOS POR LÍNEA DEL F29")
    c.font = Font(name="Arial", size=10, bold=True); c.fill = FB; c.alignment = AC
    for ci in range(1, 11):
        ws.cell(row=r, column=ci).fill = FB
    r += 2

    WHITE = "FFFFFF"
    FILL_W = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")

    orden = ["linea_1", "linea_2", "linea_5", "linea_7", "linea_9",
             "linea_10", "linea_11", "linea_12", "linea_13",
             "linea_27", "linea_28", "linea_29", "linea_31", "linea_32", "linea_33",
             "linea_60", "linea_61"]

    for lk in orden:
        if lk not in docs or not docs[lk]:
            continue

        info = LINEAS_INFO.get(lk, {})
        seccion = info.get("seccion", "")
        doc_list = docs[lk]
        columnas = _get_columnas(seccion, lk)

        header = (f"LÍNEA {info.get('linea', '?')} — {info.get('nombre', lk)} — "
                  f"Cód. {info.get('cod_cant', '?')}/{info.get('cod_monto', '?')} — "
                  f"{len(doc_list)} documento(s)")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
        color = {"debito": "E65100", "credito": "2E7D32", "retencion": "1565C0"}.get(seccion, "333333")
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        font_w = Font(name="Arial", size=9, bold=True, color=WHITE)
        for ci in range(1, 11):
            ws.cell(row=r, column=ci).fill = fill
            ws.cell(row=r, column=ci).font = font_w
        ws.cell(row=r, column=1, value=header)
        r += 1

        ws.cell(row=r, column=1, value="#").font = F8B
        ws.cell(row=r, column=1).fill = FE; ws.cell(row=r, column=1).alignment = AC
        for ci, cn in enumerate(columnas, 2):
            ws.cell(row=r, column=ci, value=cn).font = F8B
            ws.cell(row=r, column=ci).fill = FE; ws.cell(row=r, column=ci).alignment = AC
        r += 1

        first_data = r
        for i, doc in enumerate(doc_list, 1):
            values = _get_doc_values(doc, seccion, lk)
            alt = FL if i % 2 == 0 else FILL_W
            ws.cell(row=r, column=1, value=i).font = Font(name="Arial", size=7)
            ws.cell(row=r, column=1).fill = alt; ws.cell(row=r, column=1).alignment = AC
            for ci, val in enumerate(values, 2):
                cell = ws.cell(row=r, column=ci, value=val)
                cell.fill = alt
                is_monto = ci - 2 >= N_TEXT_COLS
                if is_monto and isinstance(val, (int, float)):
                    cell.number_format = NF
                    cell.font = Font(name="Arial", size=7); cell.alignment = AR
                else:
                    cell.font = Font(name="Arial", size=7); cell.alignment = AL
            r += 1
        last_data = r - 1

        ws.cell(row=r, column=1).fill = FE
        ws.cell(row=r, column=2, value="TOTAL").font = F8B
        ws.cell(row=r, column=2).fill = FE
        for ci, cn in enumerate(columnas, 2):
            ws.cell(row=r, column=ci).fill = FE
            is_monto = ci - 2 >= N_TEXT_COLS
            if is_monto:
                col_letter = get_column_letter(ci)
                ws.cell(row=r, column=ci, value=f"=SUM({col_letter}{first_data}:{col_letter}{last_data})")
                ws.cell(row=r, column=ci).number_format = NF
                ws.cell(row=r, column=ci).font = Font(name="Arial", size=7, bold=True, color="CC0000")
                ws.cell(row=r, column=ci).alignment = AR
            else:
                ws.cell(row=r, column=ci).font = F8B
        r += 1

        cod_cant = info.get("cod_cant")
        if cod_cant and codigos:
            declared = codigos.get(cod_cant, 0)
            actual = len(doc_list)
            if declared > 0 and actual < declared:
                fill_warn = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                font_warn = Font(name="Arial", size=8, bold=True, color="856404")
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
                ws.cell(row=r, column=1,
                        value=f"Se declararon {declared} documentos (cód. {cod_cant}) pero solo se identificaron {actual}. Faltan {declared - actual} documento(s).").font = font_warn
                for ci in range(1, 11):
                    ws.cell(row=r, column=ci).fill = fill_warn
                r += 1
        r += 1


# ============================================================
# Hoja 3: Alertas
# ============================================================

def _write_alertas(wb, codigos, datos):
    """Hoja 3: Alertas y notas."""
    enc = datos.get("encabezado", {})
    notas = datos.get("notas", [])
    ws = wb.create_sheet(title="Alertas y Notas")
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 80

    r = 1
    ws.cell(row=r, column=1, value="ALERTAS Y VALIDACIONES").font = Font(name="Arial", size=10, bold=True)
    ws.cell(row=r, column=1).fill = FB
    ws.cell(row=r, column=2).fill = FB
    r += 2

    def fp(v):
        if v is None or v == 0: return "$0"
        if isinstance(v, float): v = int(v)
        return f"-${abs(v):,}".replace(",", ".") if v < 0 else f"${v:,}".replace(",", ".")

    alertas = []
    if codigos.get(77, 0) > 0 and codigos.get(538, 0) > 0:
        if codigos[77] / max(codigos[538], 1) > 3:
            alertas.append(("REMANENTE", f"Remanente CF ({fp(codigos[77])}) muy superior al débito. Considerar devolución Art. 36 o Art. 27 bis."))
    if codigos.get(585, 0) > 0:
        alertas.append(("EXPORTACIÓN", f"{codigos[585]} facturas de exportación ({fp(codigos[20])} neto). Verificar calificación Aduanas."))
    if codigos.get(596, 0) > 0:
        alertas.append(("CAMBIO SUJETO", f"Retención cambio de sujeto: {fp(codigos[596])}. IVA retenido por FC de servicios digitales extranjeros."))
    if codigos.get(151, 0) > 0:
        alertas.append(("HONORARIOS", f"Retención: {fp(codigos[151])}. Tasa 2026: 15,25%."))
    if codigos.get(504, 0) > 0:
        alertas.append(("REMANENTE ANT.", f"Remanente CF arrastrado: {fp(codigos[504])}. Verificar vs código 77 del F29 anterior."))
    has_afectas = codigos.get(503, 0) > 0 or codigos.get(110, 0) > 0
    has_exentas = codigos.get(585, 0) > 0 or codigos.get(586, 0) > 0
    if has_afectas and has_exentas:
        alertas.append(("PRORRATEO", "Ventas afectas + exentas/exportación. Verificar prorrateo de CF de uso común."))

    for nota in notas:
        if isinstance(nota, tuple) and len(nota) == 2:
            alertas.append(nota)
        elif isinstance(nota, str):
            alertas.append(("NOTA", nota))

    mes = enc.get("periodo_mes", 1)
    anio = enc.get("periodo_anio", 2026)
    sig_mes = mes % 12 + 1
    sig_anio = anio if mes < 12 else anio + 1
    alertas.append(("PLAZO", f"Declarar antes del 20 de {MESES.get(sig_mes, '')} {sig_anio} (internet)."))
    alertas.append(("DISCLAIMER", "Herramienta de apoyo. NO es asesoría tributaria. Revisar con contador antes de presentar al SII."))

    fill_warn = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    for tipo, msg in alertas:
        c1 = ws.cell(row=r, column=1, value=tipo)
        c1.font = F8B
        c2 = ws.cell(row=r, column=2, value=msg)
        c2.font = Font(name="Arial", size=7)
        c2.alignment = Alignment(wrap_text=True, vertical="top")
        if tipo in ("REMANENTE", "PRORRATEO"):
            c1.fill = fill_warn; c2.fill = fill_warn
        r += 1


# ============================================================
# Función principal
# ============================================================

def generar_f29_excel(datos, output_path):
    """Genera Excel del F29 con 3 hojas."""
    codigos = calcular_f29(datos)
    enc = datos.get("encabezado", {})
    wb = openpyxl.Workbook()
    _write_f29(wb, codigos, enc)
    _write_detalle(wb, datos, codigos)
    _write_alertas(wb, codigos, datos)
    wb.save(output_path)
    return codigos


if __name__ == "__main__":
    datos_ejemplo = {
        "encabezado": {
            "rut": "78.033.706-0", "razon_social": "TOTOMENU SPA",
            "periodo_mes": 12, "periodo_anio": 2025, "folio": "8740690436",
        },
        "codigos": {
            503: 14, 502: 2541111, 538: 2541111,
            511: 2074, 514: 0, 584: 4, 562: 452579,
            519: 32, 520: 497664, 527: 1, 528: 2934, 537: 494730,
            77: 0, 89: 2046381, 48: 0, 151: 0, 153: 0,
            563: 13374273, 115: 2.5, 62: 334357, 595: 2380738,
            39: 495590, 736: 2934, 596: 492656,
            547: 2873394, 91: 2873394, 92: 0, 93: 0, 94: 2873394,
        },
    }
    output = "f29_test.xlsx"
    codigos = generar_f29_excel(datos_ejemplo, output)
    def fp(v): return f"${v:,}".replace(",", ".") if v else "$0"
    print(f"F29 generado: {output}")
    print(f"  Débito (538): {fp(codigos.get(538, 0))}")
    print(f"  Crédito (537): {fp(codigos.get(537, 0))}")
    print(f"  IVA det (89): {fp(codigos.get(89, 0))}")
    print(f"  Total (91): {fp(codigos.get(91, 0))}")
